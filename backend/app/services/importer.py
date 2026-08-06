"""Excel 导入:模板生成、逐行验证、事务性入库。

模板固定三行表头:第 1 行字段名、第 2 行说明、第 3 行示例;
导入时自第 4 行起读取数据(前三行自动跳过)。
采用“全部正确才写入”的校验方式:任一行有误即报告所有错误,数据库零写入。
"""

import io
from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.security import hash_password
from app.core.validators import is_valid_email
from app.models.assignment import AssignmentTeacher, BlockRule, CourseAssignment
from app.models.basedata import (
    ClassTrack,
    ClassUnit,
    RoomType,
    Subject,
    Teacher,
)
from app.models.period import PeriodTable
from app.models.user import Role, User, UserRole
from app.services.assignments import (
    DomainError,
    assert_within_overtime_limit,
    get_or_create_single_unit,
)

HEADER_ROWS = 3  # 字段名 + 说明 + 示例

ROOM_TYPE_BY_LABEL = {
    "普通教室": RoomType.normal,
    "专用教室": RoomType.special,
    "专科教室": RoomType.special,
    "实训场地": RoomType.workshop,
    "户外场地": RoomType.outdoor,
}
TRACK_BY_LABEL = {
    "小学": ClassTrack.elementary,
    "初中": ClassTrack.junior_high,
    "普通高中": ClassTrack.senior_high,
    "综合高中": ClassTrack.comprehensive,
    "中职": ClassTrack.vocational,
    "职业高中": ClassTrack.vocational,
}

# 每个实体的模板字段：（字段名、说明、示例）。
TEMPLATE_DEFS: dict[str, dict] = {
    "subjects": {
        "sheet": "科目",
        "columns": [
            ("名称", "必填", "数学"),
            ("领域/类别", "选填", "数学"),
            ("所需教室/场地类型", "选填：普通教室/专用教室/实训场地/户外场地", "普通教室"),
            ("默认连堂", "选填，数字 1-8，默认 1", "1"),
        ],
    },
    "teachers": {
        "sheet": "教师",
        "columns": [
            ("姓名", "必填", "王小明"),
            ("身份后四位", "选填，4 位，用于识别同名教师", "1234"),
            ("任教科目", "选填，多科以、分隔；需为已创建的科目", "数学、物理"),
            ("基本课时", "选填，数字", "20"),
            ("行政职务", "选填", "排课管理员"),
            ("行政减课", "选填，数字", "4"),
            ("外聘", "选填：是/否，默认否", "否"),
            ("登录账号", "选填，勾选创建账号时使用", "wang001"),
            ("邮箱", "选填，用于调课与代课通知", "wang@example.edu.cn"),
            ("手机号", "选填，用于联系", "13800138000"),
            ("其他联系方式", "选填", ""),
        ],
    },
    "classes": {
        "sheet": "班级",
        "columns": [
            ("年级", "必填，数字 1-12", "7"),
            ("班名", "必填", "1班"),
            ("学制", "必填：小学/初中/普通高中/综合高中/中职/职业高中", "初中"),
            ("专业/班级类别", "选填", ""),
            ("班主任", "选填，需为已创建的教师姓名", "王小明"),
            ("人数", "选填，数字", "45"),
            ("作息时间表", "选填，空白则使用学期默认作息时间表", "作息时间表（待完善）"),
        ],
    },
    "assignments": {
        "sheet": "教学任务",
        "columns": [
            ("班级", "必填，需为已创建的班名", "1班"),
            ("科目", "必填，需为已创建的科目", "数学"),
            ("教师", "必填，多位以、分隔，第一位为主讲教师", "王小明、李老师"),
            ("每周课时", "必填，数字", "5"),
            ("连堂长度", "选填，2-4；与连堂次数成对填写", "2"),
            ("连堂次数", "选填，数字", "1"),
            ("教室/场地类型", "选填：普通教室/专用教室/实训场地/户外场地", "专用教室"),
        ],
    },
}


@dataclass
class ImportResult:
    imported: int = 0
    errors: list[str] = field(default_factory=list)


def build_template(entity: str) -> bytes:
    cfg = TEMPLATE_DEFS[entity]
    wb = Workbook()
    ws = wb.active
    ws.title = cfg["sheet"]
    ws.append([c[0] for c in cfg["columns"]])
    ws.append([c[1] for c in cfg["columns"]])
    ws.append([c[2] for c in cfg["columns"]])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(row: tuple, i: int) -> str | None:
    if i >= len(row) or row[i] is None:
        return None
    text = str(row[i]).strip()
    return text or None


def _parse_int(value: str | None) -> int | None:
    if value is None:
        return None
    try:
        return int(float(value))  # 容忍 Excel 把数字读成 20.0
    except ValueError as err:
        raise ValueError(f"「{value}」不是有效数字") from err


def _data_rows(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    for idx, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True), start=HEADER_ROWS + 1
    ):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        yield idx, row


# ── 各实体导入 ────────────────────────
def _import_subjects(db: Session, semester_id: int, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    pending: list[Subject] = []
    for idx, row in _data_rows(file_bytes):
        name = _cell(row, 0)
        if not name:
            result.errors.append(f"第 {idx} 行:名称必填")
            continue
        room_label = _cell(row, 2)
        room_type = None
        if room_label:
            if room_label not in ROOM_TYPE_BY_LABEL:
                result.errors.append(f"第 {idx} 行:教室/场地类型「{room_label}」无效")
                continue
            room_type = ROOM_TYPE_BY_LABEL[room_label].value
        try:
            block = _parse_int(_cell(row, 3)) or 1
        except ValueError as e:
            result.errors.append(f"第 {idx} 行:默认连堂 {e}")
            continue
        pending.append(
            Subject(
                semester_id=semester_id, name=name, domain=_cell(row, 1),
                required_room_type=room_type, default_block_size=block,
            )
        )
    if result.errors:
        return result
    db.add_all(pending)
    db.commit()
    result.imported = len(pending)
    return result


def _import_classes(db: Session, semester_id: int, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    teachers = {
        t.name: t.id
        for t in db.scalars(select(Teacher).where(Teacher.semester_id == semester_id))
    }
    period_tables = {
        pt.name: pt.id
        for pt in db.scalars(
            select(PeriodTable).where(PeriodTable.semester_id == semester_id)
        )
    }
    # 同学期班名唯一(M6-5)。文件内重复、或与现有班级重复,都要当场说是哪一列,
    # 不能让它撞上 DB 约束变成一句看不懂的错误。
    existing_names = set(
        db.scalars(
            select(ClassUnit.name).where(ClassUnit.semester_id == semester_id)
        )
    )
    pending: list[ClassUnit] = []
    for idx, row in _data_rows(file_bytes):
        try:
            grade = _parse_int(_cell(row, 0))
        except ValueError as e:
            result.errors.append(f"第 {idx} 行:年级 {e}")
            continue
        name = _cell(row, 1)
        track_label = _cell(row, 2)
        if not grade or not name or not track_label:
            result.errors.append(f"第 {idx} 行:年级、班名、学制均必填")
            continue
        if name in existing_names:
            result.errors.append(f"第 {idx} 行:班名「{name}」在本学期重复")
            continue
        existing_names.add(name)
        if track_label not in TRACK_BY_LABEL:
            result.errors.append(f"第 {idx} 行:学制「{track_label}」无效")
            continue
        homeroom_name = _cell(row, 4)
        homeroom_id = None
        if homeroom_name:
            if homeroom_name not in teachers:
                result.errors.append(f"第 {idx} 行:班主任「{homeroom_name}」不存在")
                continue
            homeroom_id = teachers[homeroom_name]
        table_name = _cell(row, 6)
        table_id = None
        if table_name:
            if table_name not in period_tables:
                result.errors.append(f"第 {idx} 行：作息时间表“{table_name}”不存在")
                continue
            table_id = period_tables[table_name]
        try:
            count = _parse_int(_cell(row, 5))
        except ValueError as e:
            result.errors.append(f"第 {idx} 行:人数 {e}")
            continue
        pending.append(
            ClassUnit(
                semester_id=semester_id, grade=grade, name=name,
                track=TRACK_BY_LABEL[track_label].value, department=_cell(row, 3),
                student_count=count, homeroom_teacher_id=homeroom_id,
                period_table_id=table_id,
            )
        )
    if result.errors:
        return result
    db.add_all(pending)
    db.commit()
    result.imported = len(pending)
    return result


def _import_teachers(
    db: Session, semester_id: int, file_bytes: bytes, create_accounts: bool
) -> ImportResult:
    result = ImportResult()
    subjects = {
        s.name: s
        for s in db.scalars(select(Subject).where(Subject.semester_id == semester_id))
    }
    existing_keys = {
        (t.name, t.id_last4 or "")
        for t in db.scalars(select(Teacher).where(Teacher.semester_id == semester_id))
    }
    existing_usernames = set(db.scalars(select(User.username)))

    seen_keys: set[tuple[str, str]] = set()
    seen_usernames: set[str] = set()
    pending: list[tuple[Teacher, str | None]] = []  # (teacher, username or None)

    for idx, row in _data_rows(file_bytes):
        name = _cell(row, 0)
        if not name:
            result.errors.append(f"第 {idx} 行:姓名必填")
            continue
        id_last4 = _cell(row, 1)
        key = (name, id_last4 or "")
        if key in existing_keys or key in seen_keys:
            result.errors.append(f"第 {idx} 行:教师「{name}」(后四位 {id_last4 or '无'})重复")
            continue
        seen_keys.add(key)

        subject_objs: list[Subject] = []
        subj_field = _cell(row, 2)
        subj_error = False
        if subj_field:
            names = [s.strip() for s in subj_field.replace(",", "、").split("、") if s.strip()]
            for sname in names:
                if sname not in subjects:
                    result.errors.append(f"第 {idx} 行:科目「{sname}」不存在")
                    subj_error = True
                    break
                subject_objs.append(subjects[sname])
        if subj_error:
            continue
        try:
            base_periods = _parse_int(_cell(row, 3)) or 0
            admin_reduction = _parse_int(_cell(row, 5)) or 0
        except ValueError as e:
            result.errors.append(f"第 {idx} 行:{e}")
            continue
        is_external = (_cell(row, 6) or "否") == "是"

        username = _cell(row, 7)
        if create_accounts and username:
            if username in existing_usernames or username in seen_usernames:
                result.errors.append(f"第 {idx} 行:登录账号「{username}」重复")
                continue
            seen_usernames.add(username)

        email = _cell(row, 8)
        if email and not is_valid_email(email):
            result.errors.append(f"第 {idx} 行:Email「{email}」格式不正确")
            continue

        teacher = Teacher(
            semester_id=semester_id, name=name, id_last4=id_last4,
            base_periods=base_periods, admin_title=_cell(row, 4),
            admin_reduction=admin_reduction, is_external=is_external,
            email=email, phone=_cell(row, 9), line_id=_cell(row, 10),
            subjects=subject_objs,
        )
        pending.append((teacher, username if create_accounts else None))

    if result.errors:
        return result

    for teacher, username in pending:
        db.add(teacher)
        if username:
            # 绑定新建账号:teacher.user 赋值于 flush 时写入 teachers.user_id
            teacher.user = User(
                username=username,
                password_hash=hash_password(settings.default_import_password),
                display_name=teacher.name,
                must_change_password=True,
                roles=[UserRole(role=Role.teacher.value)],
            )
    db.commit()
    result.imported = len(pending)
    return result


def _import_assignments(db: Session, semester_id: int, file_bytes: bytes) -> ImportResult:
    """单班教学任务导入(班级×科目×教师×周节数,可含一组连堂)。走班群组于页面创建。"""
    result = ImportResult()
    classes = {
        c.name: c
        for c in db.scalars(select(ClassUnit).where(ClassUnit.semester_id == semester_id))
    }
    subjects = {
        s.name: s
        for s in db.scalars(select(Subject).where(Subject.semester_id == semester_id))
    }
    teachers = {
        t.name: t
        for t in db.scalars(select(Teacher).where(Teacher.semester_id == semester_id))
    }
    # (class_unit, subject, teacher_objs, periods, block, room_type)
    pending: list[tuple] = []
    for idx, row in _data_rows(file_bytes):
        class_name = _cell(row, 0)
        subj_name = _cell(row, 1)
        teacher_field = _cell(row, 2)
        if not class_name or not subj_name or not teacher_field:
            result.errors.append(f"第 {idx} 行:班级、科目、教师均必填")
            continue
        if class_name not in classes:
            result.errors.append(f"第 {idx} 行:班级「{class_name}」不存在")
            continue
        if subj_name not in subjects:
            result.errors.append(f"第 {idx} 行:科目「{subj_name}」不存在")
            continue
        names = [s.strip() for s in teacher_field.replace(",", "、").split("、") if s.strip()]
        teacher_objs = []
        t_error = False
        for tname in names:
            if tname not in teachers:
                result.errors.append(f"第 {idx} 行:教师「{tname}」不存在")
                t_error = True
                break
            teacher_objs.append(teachers[tname])
        if t_error:
            continue
        try:
            periods = _parse_int(_cell(row, 3))
        except ValueError as e:
            result.errors.append(f"第 {idx} 行:每周节数 {e}")
            continue
        if not periods or periods < 1:
            result.errors.append(f"第 {idx} 行:每周节数必填且需大于 0")
            continue
        # 连堂(选填,长度与次数需成对)
        try:
            block_size = _parse_int(_cell(row, 4))
            block_count = _parse_int(_cell(row, 5))
        except ValueError as e:
            result.errors.append(f"第 {idx} 行:连堂 {e}")
            continue
        block: tuple[int, int] | None = None
        if block_size is not None or block_count is not None:
            if block_size is None or block_count is None:
                result.errors.append(f"第 {idx} 行:连堂长度与连堂次数需成对填写")
                continue
            if not 2 <= block_size <= 4 or block_count < 1:
                result.errors.append(f"第 {idx} 行:连堂长度需为 2-4、次数需大于 0")
                continue
            if block_size * block_count > periods:
                result.errors.append(f"第 {idx} 行:连堂总节数超过每周节数")
                continue
            block = (block_size, block_count)
        room_label = _cell(row, 6)
        room_type = None
        if room_label:
            if room_label not in ROOM_TYPE_BY_LABEL:
                result.errors.append(f"第 {idx} 行:教室/场地类型「{room_label}」无效")
                continue
            room_type = ROOM_TYPE_BY_LABEL[room_label].value
        pending.append((classes[class_name], subjects[subj_name], teacher_objs, periods, block,
                        room_type))

    if result.errors:
        return result

    for class_unit, subject, teacher_objs, periods, block, room_type in pending:
        unit = get_or_create_single_unit(db, class_unit)
        assignment = CourseAssignment(
            semester_id=semester_id, scheduling_unit_id=unit.id, subject_id=subject.id,
            periods_per_week=periods, required_room_type=room_type,
        )
        db.add(assignment)
        db.flush()
        for i, t in enumerate(teacher_objs):
            assignment.teachers.append(
                AssignmentTeacher(teacher_id=t.id, is_lead=(i == 0))
            )
        if block is not None:
            assignment.block_rules.append(
                BlockRule(block_size=block[0], count_per_week=block[1])
            )

    # 超课时上限:整批一起检核。汇入是「一次几百笔」的操作,
    # 逐列阻止只会让用户改一次汇一次;这里一次把所有超标的人列出来。
    db.flush()
    touched = {t.id for _, _, teacher_objs, _, _, _ in pending for t in teacher_objs}
    try:
        assert_within_overtime_limit(db, semester_id, touched)
    except DomainError as exc:
        db.rollback()
        result.errors.append(exc.message)
        result.imported = 0
        return result

    db.commit()
    result.imported = len(pending)
    return result


def run_import(
    db: Session, entity: str, semester_id: int, file_bytes: bytes, create_accounts: bool = False
) -> ImportResult:
    if entity == "subjects":
        return _import_subjects(db, semester_id, file_bytes)
    if entity == "classes":
        return _import_classes(db, semester_id, file_bytes)
    if entity == "teachers":
        return _import_teachers(db, semester_id, file_bytes, create_accounts)
    if entity == "assignments":
        return _import_assignments(db, semester_id, file_bytes)
    raise ValueError(f"未知的导入类型:{entity}")
