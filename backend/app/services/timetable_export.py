"""课表导出:把已发布课表摊成格线,生成 Excel / PDF(HTML)/PNG(M5-1)。

三种对象(班级 / 教师 / 教室/场地)共用同一个格线模型 `Grid`,再交给各格式的渲染器,
确保三种格式与页面课表内容一致(验收①)。数据来源统一是**已发布**课表(D4 快照)。

- Excel 由 openpyxl 生成,可在 api 同步跑(轻量)。
- PDF 需 WeasyPrint(系统依赖+中文字体只在 worker),故 PDF/PNG 走 worker 背景渲染;
  本模块只负责产出 HTML,实际 write_pdf 在 `app.workers.export_job`。
"""

import io
import zipfile
from dataclasses import dataclass, field
from datetime import date

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.basedata import ClassUnit, Room, Teacher
from app.models.period import Period, PeriodTable, PeriodType
from app.models.semester import Semester
from app.models.timetable import ScheduleEntry, Timetable, TimetableStatus
from app.services import period_tables as pt_service
from app.services import school_rules


class ExportError(Exception):
    """导出前置条件不成立(调用方转为 4xx)。"""


# ── 格线模型 ────────────────────────────────────────────────
@dataclass(frozen=True, slots=True)
class Cell:
    lines: tuple[str, ...] = ()  # 科目 / 教师或班级 / 教室(逐行)
    span: int = 1  # 连堂占几列
    covered: bool = False  # 被上方连堂覆盖,不绘


@dataclass
class Row:
    period_no: int
    label: str
    is_regular: bool
    cells: list[Cell] = field(default_factory=list)  # 长度 = num_weekdays


@dataclass
class Grid:
    title: str
    num_weekdays: int
    rows: list[Row] = field(default_factory=list)

    @property
    def weekday_names(self) -> list[str]:
        return list(school_rules.weekday_names()[: self.num_weekdays])


@dataclass(frozen=True, slots=True)
class Meta:
    school_name: str
    semester_label: str
    timetable_name: str
    printed_on: date


@dataclass(frozen=True, slots=True)
class _EntryView:
    weekday: int
    period_no: int
    span: int
    subject: str
    teachers: str
    classes: str
    room: str
    teacher_ids: frozenset[int]
    class_ids: frozenset[int]
    room_id: int | None


class _Published:
    """一份已发布课表的导出来源:一次查询单元格、作息时间表和对象列表。"""

    def __init__(self, db: Session, semester_id: int) -> None:
        self.db = db
        semester = db.get(Semester, semester_id)
        if semester is None:
            raise ExportError("找不到学期")
        timetable = db.scalar(
            select(Timetable).where(
                Timetable.semester_id == semester_id,
                Timetable.status == TimetableStatus.published.value,
            )
        )
        if timetable is None:
            raise ExportError("此学期尚无已发布的课表")
        self.semester: Semester = semester
        self.timetable: Timetable = timetable
        self.entries = [
            _view(e)
            for e in db.scalars(
                select(ScheduleEntry).where(ScheduleEntry.timetable_id == self.timetable.id)
            )
        ]
        self.tables = {
            t.id: t
            for t in db.scalars(select(PeriodTable).where(PeriodTable.semester_id == semester_id))
        }

    def meta(self) -> Meta:
        from app.core import clock

        return Meta(
            school_name=settings.school_name,
            semester_label=self.semester.label,
            timetable_name=self.timetable.name,
            printed_on=clock.school_today(),
        )

    def default_table(self) -> PeriodTable | None:
        tables = list(self.tables.values())
        return next((t for t in tables if t.is_default), tables[0] if tables else None)

    def class_table(self, cls: ClassUnit) -> PeriodTable | None:
        t = pt_service.resolve_period_table(self.db, cls)
        return t or self.default_table()


def _view(e: ScheduleEntry) -> _EntryView:
    a = e.assignment
    su = a.scheduling_unit
    room = e.room if e.room is not None else a.room
    return _EntryView(
        weekday=e.weekday,
        period_no=e.period_no,
        span=e.span,
        subject=a.subject.name,
        teachers="、".join(at.teacher.name for at in a.teachers),
        classes="、".join(m.class_unit.name for m in su.members),
        room=room.name if room else "",
        teacher_ids=frozenset(at.teacher_id for at in a.teachers),
        class_ids=frozenset(m.class_unit_id for m in su.members),
        room_id=e.effective_room_id,
    )


def _grid_from(
    table: PeriodTable | None,
    title: str,
    entries: list[_EntryView],
    lines_of,
) -> Grid:
    """把某对象的单元格排进作息时间表格线。`lines_of(entry)` 决定每格显示哪几行。"""
    if table is None:
        return Grid(title=title, num_weekdays=5, rows=[])
    num_weekdays = table.num_weekdays
    periods = sorted(table.periods, key=lambda p: (p.period_no, p.weekday))
    # 每个 period_no 取一个代表(名称/类型),weekday 小者优先
    by_no: dict[int, Period] = {}
    for p in periods:
        by_no.setdefault(p.period_no, p)
    order = sorted(by_no)

    # (weekday, period_no) → entry
    placed: dict[tuple[int, int], _EntryView] = {(e.weekday, e.period_no): e for e in entries}
    covered: set[tuple[int, int]] = set()
    for e in entries:
        for k in range(1, e.span):
            covered.add((e.weekday, e.period_no + k))

    grid = Grid(title=title, num_weekdays=num_weekdays)
    for pno in order:
        rep = by_no[pno]
        is_regular = rep.type == PeriodType.regular.value
        row = Row(period_no=pno, label=rep.name, is_regular=is_regular)
        for wd in range(1, num_weekdays + 1):
            if (wd, pno) in covered:
                row.cells.append(Cell(covered=True))
            elif (wd, pno) in placed:
                e = placed[(wd, pno)]
                row.cells.append(Cell(lines=tuple(lines_of(e)), span=e.span))
            else:
                row.cells.append(Cell())
        grid.rows.append(row)
    return grid


# ── 三种对象 → Grid ─────────────────────────────────────────
def _class_lines(e: _EntryView) -> list[str]:
    return [x for x in (e.subject, e.teachers, e.room) if x]


def _teacher_lines(e: _EntryView) -> list[str]:
    return [x for x in (e.subject, e.classes, e.room) if x]


def _room_lines(e: _EntryView) -> list[str]:
    return [x for x in (e.subject, e.classes, e.teachers) if x]


def class_grid(pub: _Published, cls: ClassUnit) -> Grid:
    entries = [e for e in pub.entries if cls.id in e.class_ids]
    return _grid_from(
        pub.class_table(cls),
        f"{cls.grade}年{cls.name} {school_rules.export_label('timetable')}",
        entries,
        _class_lines,
    )


def teacher_grid(pub: _Published, teacher: Teacher) -> Grid:
    entries = [e for e in pub.entries if teacher.id in e.teacher_ids]
    return _grid_from(
        pub.default_table(),
        f"{teacher.name} {school_rules.export_label('timetable')}",
        entries,
        _teacher_lines,
    )


def room_grid(pub: _Published, room: Room) -> Grid:
    entries = [e for e in pub.entries if e.room_id == room.id]
    return _grid_from(
        pub.default_table(),
        f"{room.name} {school_rules.export_label('timetable')}",
        entries,
        _room_lines,
    )


def build_grid(db: Session, semester_id: int, view: str, target_id: int) -> tuple[Grid, Meta]:
    pub = _Published(db, semester_id)
    if view == "class":
        obj = db.get(ClassUnit, target_id)
        if obj is None or obj.semester_id != semester_id:
            raise ExportError("找不到班级")
        return class_grid(pub, obj), pub.meta()
    if view == "teacher":
        obj_t = db.get(Teacher, target_id)
        if obj_t is None or obj_t.semester_id != semester_id:
            raise ExportError("找不到教师")
        return teacher_grid(pub, obj_t), pub.meta()
    if view == "room":
        obj_r = db.get(Room, target_id)
        if obj_r is None or obj_r.semester_id != semester_id:
            raise ExportError("找不到教室/场地")
        return room_grid(pub, obj_r), pub.meta()
    raise ExportError(f"未知的查看类型:{view}")


def _classes(db: Session, semester_id: int) -> list[ClassUnit]:
    return list(
        db.scalars(
            select(ClassUnit)
            .where(ClassUnit.semester_id == semester_id)
            .order_by(ClassUnit.grade, ClassUnit.name)
        )
    )


def school_workbook(db: Session, semester_id: int) -> bytes:
    """全校总表:一个 Excel 文件,每班一个工作表。"""
    pub = _Published(db, semester_id)
    grids = [class_grid(pub, c) for c in _classes(db, semester_id)]
    return grids_to_xlsx(grids, pub.meta())


def class_batch_zip(db: Session, semester_id: int) -> bytes:
    """批量导出:全部班级各一个 Excel 文件,打包成 zip。"""
    pub = _Published(db, semester_id)
    meta = pub.meta()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for c in _classes(db, semester_id):
            grid = class_grid(pub, c)
            z.writestr(f"{c.grade}年{c.name}.xlsx", grids_to_xlsx([grid], meta))
    return buf.getvalue()


# ── Excel 渲染 ──────────────────────────────────────────────
def _safe_sheet_title(title: str, used: set[str]) -> str:
    # Excel 工作表名称 ≤31 字,且不可包含 : \ / ? * [ ]
    clean = title
    for ch in ":\\/?*[]":
        clean = clean.replace(ch, " ")
    clean = clean[:28].strip() or school_rules.export_label("timetable")
    name, i = clean, 1
    while name in used:
        name = f"{clean[:26]}~{i}"
        i += 1
    used.add(name)
    return name


def grids_to_xlsx(grids: list[Grid], meta: Meta) -> bytes:
    """每张 Grid 对应一个工作表。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    head_fill = PatternFill("solid", fgColor="E8E8E8")
    shade = PatternFill("solid", fgColor="F5F5F5")
    used: set[str] = set()

    for grid in grids:
        ws = wb.create_sheet(_safe_sheet_title(grid.title, used))
        ws.append([grid.title])
        ws.append(
            [
                f"{meta.school_name}　{meta.semester_label}　"
                f"{school_rules.export_label('printed_on')}：{meta.printed_on}"
            ]
        )
        ws["A1"].font = Font(bold=True, size=14)
        header = [school_rules.export_label("period"), *grid.weekday_names]
        ws.append(header)
        head_row = ws.max_row
        for c in ws[head_row]:
            c.font = Font(bold=True)
            c.fill = head_fill
            c.alignment = center
            c.border = border

        for row in grid.rows:
            values = [row.label]
            for cell in row.cells:
                values.append("\n".join(cell.lines) if cell.lines else "")
            ws.append(values)
            r = ws.max_row
            for ci, c in enumerate(ws[r]):
                c.alignment = center
                c.border = border
                if ci == 0 or not row.is_regular:
                    c.fill = shade
            # 连堂:垂直合并
            for ci, cell in enumerate(row.cells):
                if cell.span > 1:
                    col = ci + 2
                    ws.merge_cells(
                        start_row=r, start_column=col, end_row=r + cell.span - 1, end_column=col
                    )

        ws.column_dimensions["A"].width = 10
        for col in range(2, 2 + grid.num_weekdays):
            ws.column_dimensions[ws.cell(row=head_row, column=col).column_letter].width = 16
        ws.freeze_panes = ws.cell(row=head_row + 1, column=2)

    if not wb.sheetnames:  # 全空:给一张空白页避免坏档
        wb.create_sheet(school_rules.export_label("timetable"))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── HTML 渲染(供 worker 转 PDF/PNG)─────────────────────────
def _esc(s: str) -> str:
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def grid_to_html(grid: Grid, meta: Meta) -> str:
    """单一对象的 A4 纵向课表 HTML。中文由 worker 镜像内嵌的 Noto CJK 呈现。"""
    head_cells = "".join(f"<th>{_esc(w)}</th>" for w in grid.weekday_names)
    period_label = _esc(school_rules.export_label("period"))
    subtitle = (
        f"{_esc(meta.school_name)}　{_esc(meta.semester_label)}　"
        f"{_esc(meta.timetable_name)}　"
        f"{_esc(school_rules.export_label('printed_on'))}：{meta.printed_on}"
    )
    font_stack = '"Noto Sans CJK SC", "Noto Sans SC", sans-serif'
    body_rows = []
    for row in grid.rows:
        tds = [f'<th class="pno">{_esc(row.label)}</th>']
        for cell in row.cells:
            if cell.covered:
                continue
            cls = "cell" if row.is_regular else "cell rest"
            span = f' rowspan="{cell.span}"' if cell.span > 1 else ""
            inner = "<br>".join(_esc(x) for x in cell.lines)
            tds.append(f'<td class="{cls}"{span}>{inner}</td>')
        body_rows.append(f"<tr>{''.join(tds)}</tr>")
    return f"""<!doctype html><html><head><meta charset="utf-8"><style>
@page {{ size: A4 portrait; margin: 14mm; }}
* {{ font-family: {font_stack}; }}
h1 {{ font-size: 18px; text-align: center; margin: 0 0 4px; }}
.meta {{ text-align: center; font-size: 12px; color: #444; margin-bottom: 10px; }}
table {{ border-collapse: collapse; width: 100%; table-layout: fixed; }}
th, td {{ border: 1px solid #333; padding: 4px; text-align: center; font-size: 12px;
  vertical-align: middle; word-break: break-all; }}
thead th {{ background: #e8e8e8; }}
.pno {{ background: #f2f2f2; width: 60px; font-size: 11px; }}
td.cell {{ height: 46px; }}
td.rest {{ background: #f7f7f7; color: #888; }}
</style></head><body>
<h1>{_esc(grid.title)}</h1>
<div class="meta">{subtitle}</div>
<table><thead><tr><th class="pno">{period_label}</th>{head_cells}</tr></thead>
<tbody>{"".join(body_rows)}</tbody></table>
</body></html>"""
