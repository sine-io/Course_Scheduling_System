"""初始基础数据组合导入：解析、预览与单事务提交。"""

from __future__ import annotations

import hashlib
import io
import json
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from typing import Any, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.validators import is_valid_email
from app.models.basedata import ClassUnit, Room, Subject, Teacher
from app.services.importer import ROOM_TYPE_BY_LABEL, TRACK_BY_LABEL

HEADER_ROWS = 3
PLAN_VERSION = 1
Status = Literal["new", "unchanged", "changed", "conflict"]

SHEET_DEFS: dict[str, dict[str, Any]] = {
    "subjects": {
        "label": "科目",
        "columns": [
            ("名称", "必填；同一学期用于识别科目", "数学"),
            ("领域/类别", "选填", "数学"),
            (
                "所需教室/场地类型",
                "选填：普通教室/专用教室/实训场地/户外场地",
                "普通教室",
            ),
            ("默认连堂", "选填，数字 1-8，默认 1", 1),
            ("主科", "选填：是/否，默认否", "是"),
        ],
    },
    "teachers": {
        "label": "教师",
        "columns": [
            ("姓名", "必填；同名教师请填写不同的身份后四位", "王老师"),
            ("身份后四位", "选填，4 位数字，用于区分同名教师", "1234"),
            ("任教科目", "选填，多科以顿号分隔；可引用本文件中的科目", "数学、物理"),
            ("基本课时", "选填，非负整数，默认 0", 18),
            ("行政职务", "选填", "年级负责人"),
            ("行政减课", "选填，非负整数，默认 0", 2),
            ("外聘", "选填：是/否，默认否", "否"),
            ("邮箱", "选填", "teacher@example.edu.cn"),
            ("手机号", "选填", "13800138000"),
            ("其他联系方式", "选填", ""),
        ],
    },
    "classes": {
        "label": "班级",
        "columns": [
            ("年级", "必填，数字 1-12", 7),
            ("班名", "必填；同一学期用于识别班级", "七年级1班"),
            ("学制", "必填：小学/初中/普通高中/综合高中/中职/职业高中", "初中"),
            ("专业/班级类别", "选填", ""),
            ("班主任", "选填；可引用本文件中的教师姓名", "王老师"),
            ("人数", "选填，非负整数", 42),
        ],
    },
    "rooms": {
        "label": "教室",
        "columns": [
            ("名称", "必填；同一学期用于识别教室/场地", "实验室A"),
            ("类型", "必填：普通教室/专用教室/实训场地/户外场地", "专用教室"),
            ("容量", "选填，非负整数", 48),
            ("适用科目", "选填，多科以顿号分隔；可引用本文件中的科目", "物理、化学"),
        ],
    },
}

ROOM_LABEL_BY_TYPE = {
    "normal": "普通教室",
    "special": "专用教室",
    "workshop": "实训场地",
    "outdoor": "户外场地",
}
TRACK_LABEL_BY_TYPE = {
    "elementary": "小学",
    "junior_high": "初中",
    "senior_high": "普通高中",
    "comprehensive": "综合高中",
    "vocational": "中职",
}


def _room_label(value: str | None) -> str | None:
    return ROOM_LABEL_BY_TYPE.get(value, value) if value is not None else None


class InvalidWorkbookError(ValueError):
    """上传内容不是可读取的 xlsx。"""


@dataclass(slots=True)
class LocatedError:
    sheet: str
    row: int
    field: str
    message: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "row": self.row,
            "field": self.field,
            "message": self.message,
        }


@dataclass(slots=True)
class PlannedRow:
    entity: str
    sheet: str
    row: int
    identity: str
    values: dict[str, Any] = field(default_factory=dict)
    status: Status = "new"
    changes: list[dict[str, Any]] = field(default_factory=list)
    errors: list[LocatedError] = field(default_factory=list)
    existing: Subject | Teacher | ClassUnit | Room | None = None
    applied: Subject | Teacher | ClassUnit | Room | None = None

    def add_error(self, field_name: str, message: str) -> None:
        error = LocatedError(self.sheet, self.row, field_name, message)
        if error not in self.errors:
            self.errors.append(error)
        self.status = "conflict"
        self.changes = []

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "row": self.row,
            "identity": self.identity,
            "status": self.status,
            "changes": self.changes,
            "errors": [error.as_dict() for error in self.errors],
        }


@dataclass(slots=True)
class ImportPlan:
    semester_id: int
    fingerprint: str
    rows: dict[str, list[PlannedRow]]

    @property
    def errors(self) -> list[LocatedError]:
        return [error for rows in self.rows.values() for row in rows for error in row.errors]

    def as_dict(self) -> dict[str, Any]:
        counts: dict[str, int] = {
            "new": 0,
            "unchanged": 0,
            "changed": 0,
            "conflict": 0,
        }
        for rows in self.rows.values():
            for row in rows:
                counts[row.status] += 1
        return {
            "fingerprint": self.fingerprint,
            "can_commit": counts["conflict"] == 0,
            "has_changes": counts["changed"] > 0,
            "counts": counts,
            "sheets": [
                {
                    "key": entity,
                    "label": definition["label"],
                    "rows": [row.as_dict() for row in self.rows[entity]],
                }
                for entity, definition in SHEET_DEFS.items()
            ],
            "errors": [error.as_dict() for error in self.errors],
        }


def build_template() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for definition in SHEET_DEFS.values():
        sheet = workbook.create_sheet(definition["label"])
        sheet.append([column[0] for column in definition["columns"]])
        sheet.append([column[1] for column in definition["columns"]])
        sheet.append([column[2] for column in definition["columns"]])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="37624F")
        sheet.freeze_panes = "A4"
        sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(definition['columns'])).coordinate}"
        for column in sheet.columns:
            longest = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(longest + 2, 14), 38)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _text(row: tuple[Any, ...], index: int) -> str | None:
    if index >= len(row) or row[index] is None:
        return None
    value = row[index]
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def _integer(row: tuple[Any, ...], index: int) -> int | None:
    value = _text(row, index)
    if value is None:
        return None
    try:
        number = float(value)
    except ValueError as exc:
        raise ValueError(f"「{value}」不是有效整数") from exc
    if not number.is_integer():
        raise ValueError(f"「{value}」不是有效整数")
    return int(number)


def _boolean(row: tuple[Any, ...], index: int, *, default: bool = False) -> bool:
    value = _text(row, index)
    if value is None:
        return default
    normalized = value.lower()
    if normalized in {"是", "true", "1", "yes"}:
        return True
    if normalized in {"否", "false", "0", "no"}:
        return False
    raise ValueError(f"「{value}」无效，请填写是或否")


def _names(value: str | None) -> tuple[str, ...]:
    if value is None:
        return ()
    names = [item.strip() for item in value.replace(",", "、").split("、")]
    return tuple(dict.fromkeys(item for item in names if item))


def _workbook_rows(
    file_bytes: bytes,
) -> tuple[dict[str, list[tuple[int, tuple[Any, ...]]]], list[PlannedRow]]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise InvalidWorkbookError("无法读取文件，请确认为有效的 Excel 文件") from exc

    result: dict[str, list[tuple[int, tuple[Any, ...]]]] = {}
    structural_rows: list[PlannedRow] = []
    for entity, definition in SHEET_DEFS.items():
        sheet_name = definition["label"]
        expected = [column[0] for column in definition["columns"]]
        if sheet_name not in workbook.sheetnames:
            row = PlannedRow(entity, sheet_name, 1, "工作表设置")
            row.add_error("工作表", f"缺少「{sheet_name}」工作表")
            structural_rows.append(row)
            result[entity] = []
            continue
        sheet = workbook[sheet_name]
        actual = [cell.value for cell in sheet[1]][: len(expected)]
        if actual != expected:
            row = PlannedRow(entity, sheet_name, 1, "工作表设置")
            row.add_error("表头", f"表头应为：{'、'.join(expected)}")
            structural_rows.append(row)
            result[entity] = []
            continue
        rows: list[tuple[int, tuple[Any, ...]]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=HEADER_ROWS + 1, values_only=True),
            start=HEADER_ROWS + 1,
        ):
            if all(value is None or str(value).strip() == "" for value in values):
                continue
            rows.append((row_number, values))
        result[entity] = rows
    workbook.close()
    return result, structural_rows


def _parse_subjects(source: list[tuple[int, tuple[Any, ...]]]) -> list[PlannedRow]:
    rows: list[PlannedRow] = []
    for row_number, raw in source:
        name = _text(raw, 0)
        row = PlannedRow("subjects", "科目", row_number, name or f"第 {row_number} 行")
        if not name:
            row.add_error("名称", "名称必填")
        room_label = _text(raw, 2)
        room_type = None
        if room_label:
            room = ROOM_TYPE_BY_LABEL.get(room_label)
            if room is None:
                row.add_error("所需教室/场地类型", f"类型「{room_label}」无效")
            else:
                room_type = room.value
        try:
            block_size = _integer(raw, 3) or 1
            if not 1 <= block_size <= 8:
                row.add_error("默认连堂", "默认连堂须在 1 至 8 之间")
        except ValueError as exc:
            block_size = 1
            row.add_error("默认连堂", str(exc))
        try:
            is_major = _boolean(raw, 4)
        except ValueError as exc:
            is_major = False
            row.add_error("主科", str(exc))
        row.values = {
            "name": name,
            "domain": _text(raw, 1),
            "required_room_type": room_type,
            "default_block_size": block_size,
            "is_major": is_major,
        }
        rows.append(row)
    _flag_duplicate_names(rows, "名称", "文件中科目名称重复")
    return rows


def _parse_teachers(source: list[tuple[int, tuple[Any, ...]]]) -> list[PlannedRow]:
    rows: list[PlannedRow] = []
    for row_number, raw in source:
        name = _text(raw, 0)
        id_last4 = _text(raw, 1)
        identity = name or f"第 {row_number} 行"
        if name and id_last4:
            identity = f"{name}（{id_last4}）"
        row = PlannedRow("teachers", "教师", row_number, identity)
        if not name:
            row.add_error("姓名", "姓名必填")
        if id_last4 and (len(id_last4) != 4 or not id_last4.isdigit()):
            row.add_error("身份后四位", "身份后四位须为 4 位数字")
        try:
            base_periods = _integer(raw, 3) or 0
            if base_periods < 0:
                row.add_error("基本课时", "基本课时不可为负数")
        except ValueError as exc:
            base_periods = 0
            row.add_error("基本课时", str(exc))
        try:
            admin_reduction = _integer(raw, 5) or 0
            if admin_reduction < 0:
                row.add_error("行政减课", "行政减课不可为负数")
        except ValueError as exc:
            admin_reduction = 0
            row.add_error("行政减课", str(exc))
        try:
            is_external = _boolean(raw, 6)
        except ValueError as exc:
            is_external = False
            row.add_error("外聘", str(exc))
        email = _text(raw, 7)
        if email and not is_valid_email(email):
            row.add_error("邮箱", f"邮箱「{email}」格式不正确")
        row.values = {
            "name": name,
            "id_last4": id_last4,
            "subject_names": _names(_text(raw, 2)),
            "base_periods": base_periods,
            "admin_title": _text(raw, 4),
            "admin_reduction": admin_reduction,
            "is_external": is_external,
            "email": email,
            "phone": _text(raw, 8),
            "line_id": _text(raw, 9),
        }
        rows.append(row)

    exact_counts = Counter(
        (row.values.get("name"), row.values.get("id_last4"))
        for row in rows
        if row.values.get("name")
    )
    same_name = Counter(row.values.get("name") for row in rows if row.values.get("name"))
    for row in rows:
        name = row.values.get("name")
        key = (name, row.values.get("id_last4"))
        if name and exact_counts[key] > 1:
            row.add_error("姓名", "文件中教师姓名与身份后四位重复")
        elif name and not row.values.get("id_last4") and same_name[name] > 1:
            row.add_error("身份后四位", "文件中有多位同名教师，请填写身份后四位")
    return rows


def _parse_classes(source: list[tuple[int, tuple[Any, ...]]]) -> list[PlannedRow]:
    rows: list[PlannedRow] = []
    for row_number, raw in source:
        name = _text(raw, 1)
        row = PlannedRow("classes", "班级", row_number, name or f"第 {row_number} 行")
        try:
            grade = _integer(raw, 0)
            if grade is None or not 1 <= grade <= 12:
                row.add_error("年级", "年级须在 1 至 12 之间")
        except ValueError as exc:
            grade = None
            row.add_error("年级", str(exc))
        if not name:
            row.add_error("班名", "班名必填")
        track_label = _text(raw, 2)
        track = TRACK_BY_LABEL.get(track_label or "")
        if track is None:
            row.add_error("学制", f"学制「{track_label or ''}」无效")
        try:
            student_count = _integer(raw, 5)
            if student_count is not None and student_count < 0:
                row.add_error("人数", "人数不可为负数")
        except ValueError as exc:
            student_count = None
            row.add_error("人数", str(exc))
        row.values = {
            "grade": grade,
            "name": name,
            "track": track.value if track else None,
            "department": _text(raw, 3),
            "homeroom_name": _text(raw, 4),
            "student_count": student_count,
        }
        rows.append(row)
    _flag_duplicate_names(rows, "班名", "文件中班名重复")
    return rows


def _parse_rooms(source: list[tuple[int, tuple[Any, ...]]]) -> list[PlannedRow]:
    rows: list[PlannedRow] = []
    for row_number, raw in source:
        name = _text(raw, 0)
        row = PlannedRow("rooms", "教室", row_number, name or f"第 {row_number} 行")
        if not name:
            row.add_error("名称", "名称必填")
        room_label = _text(raw, 1)
        room = ROOM_TYPE_BY_LABEL.get(room_label or "")
        if room is None:
            row.add_error("类型", f"类型「{room_label or ''}」无效")
        try:
            capacity = _integer(raw, 2)
            if capacity is not None and capacity < 0:
                row.add_error("容量", "容量不可为负数")
        except ValueError as exc:
            capacity = None
            row.add_error("容量", str(exc))
        row.values = {
            "name": name,
            "room_type": room.value if room else None,
            "capacity": capacity,
            "subject_names": _names(_text(raw, 3)),
        }
        rows.append(row)
    _flag_duplicate_names(rows, "名称", "文件中教室/场地名称重复")
    return rows


def _flag_duplicate_names(rows: list[PlannedRow], field_name: str, message: str) -> None:
    counts = Counter(row.values.get("name") for row in rows if row.values.get("name"))
    for row in rows:
        if row.values.get("name") and counts[row.values["name"]] > 1:
            row.add_error(field_name, message)


def _add_change(
    changes: list[dict[str, Any]],
    field_name: str,
    before: Any,
    after: Any,
) -> None:
    if before != after:
        changes.append({"field": field_name, "before": before, "after": after})


def _classify_subjects(db: Session, semester_id: int, rows: list[PlannedRow]) -> list[Subject]:
    existing = list(
        db.scalars(select(Subject).where(Subject.semester_id == semester_id)).all()
    )
    by_name: dict[str, list[Subject]] = defaultdict(list)
    for subject in existing:
        by_name[subject.name].append(subject)
    for row in rows:
        if row.errors:
            continue
        matches = by_name[row.values["name"]]
        if len(matches) > 1:
            row.add_error("名称", "现有数据中有多个同名科目，无法确定更新对象")
            continue
        if not matches:
            row.status = "new"
            continue
        subject = matches[0]
        row.existing = subject
        changes: list[dict[str, Any]] = []
        _add_change(changes, "领域/类别", subject.domain, row.values["domain"])
        _add_change(
            changes,
            "所需教室/场地类型",
            _room_label(subject.required_room_type),
            _room_label(row.values["required_room_type"]),
        )
        _add_change(
            changes, "默认连堂", subject.default_block_size, row.values["default_block_size"]
        )
        _add_change(changes, "主科", subject.is_major, row.values["is_major"])
        row.changes = changes
        row.status = "changed" if changes else "unchanged"
    return existing


def _subject_candidates(
    existing: list[Subject], rows: list[PlannedRow]
) -> dict[str, list[Subject | PlannedRow]]:
    candidates: dict[str, list[Subject | PlannedRow]] = defaultdict(list)
    for subject in existing:
        candidates[subject.name].append(subject)
    for row in rows:
        if not row.errors and row.existing is None:
            candidates[row.values["name"]].append(row)
    return candidates


def _validate_subject_references(
    row: PlannedRow,
    field_name: str,
    names: tuple[str, ...],
    candidates: dict[str, list[Subject | PlannedRow]],
) -> None:
    for name in names:
        matches = candidates.get(name, [])
        if not matches:
            row.add_error(field_name, f"科目「{name}」不存在")
        elif len(matches) > 1:
            row.add_error(field_name, f"科目「{name}」有多个同名记录，无法确定引用")


def _classify_teachers(
    db: Session,
    semester_id: int,
    rows: list[PlannedRow],
    subject_candidates: dict[str, list[Subject | PlannedRow]],
) -> list[Teacher]:
    existing = list(
        db.scalars(select(Teacher).where(Teacher.semester_id == semester_id)).all()
    )
    by_name: dict[str, list[Teacher]] = defaultdict(list)
    by_exact: dict[tuple[str, str], list[Teacher]] = defaultdict(list)
    for teacher in existing:
        by_name[teacher.name].append(teacher)
        by_exact[(teacher.name, teacher.id_last4 or "")].append(teacher)

    for row in rows:
        if row.errors:
            continue
        _validate_subject_references(
            row, "任教科目", row.values["subject_names"], subject_candidates
        )
        if row.errors:
            continue
        name = row.values["name"]
        id_last4 = row.values["id_last4"]
        matches = by_exact[(name, id_last4)] if id_last4 else by_name[name]
        if len(matches) > 1:
            message = (
                "现有数据中有多位教师使用相同姓名和身份后四位"
                if id_last4
                else "现有数据中有多位同名教师，请填写身份后四位以明确对应关系"
            )
            row.add_error("身份后四位", message)
            continue
        if not matches:
            row.status = "new"
            continue
        teacher = matches[0]
        row.existing = teacher
        changes: list[dict[str, Any]] = []
        before_subjects = sorted(subject.name for subject in teacher.subjects)
        after_subjects = sorted(row.values["subject_names"])
        _add_change(changes, "任教科目", before_subjects, after_subjects)
        _add_change(changes, "基本课时", teacher.base_periods, row.values["base_periods"])
        _add_change(changes, "行政职务", teacher.admin_title, row.values["admin_title"])
        _add_change(changes, "行政减课", teacher.admin_reduction, row.values["admin_reduction"])
        _add_change(changes, "外聘", teacher.is_external, row.values["is_external"])
        _add_change(changes, "邮箱", teacher.email, row.values["email"])
        _add_change(changes, "手机号", teacher.phone, row.values["phone"])
        _add_change(changes, "其他联系方式", teacher.line_id, row.values["line_id"])
        row.changes = changes
        row.status = "changed" if changes else "unchanged"
    return existing


def _teacher_candidates(
    existing: list[Teacher], rows: list[PlannedRow]
) -> dict[str, list[Teacher | PlannedRow]]:
    candidates: dict[str, list[Teacher | PlannedRow]] = defaultdict(list)
    for teacher in existing:
        candidates[teacher.name].append(teacher)
    for row in rows:
        if not row.errors and row.existing is None:
            candidates[row.values["name"]].append(row)
    return candidates


def _classify_classes(
    db: Session,
    semester_id: int,
    rows: list[PlannedRow],
    teacher_candidates: dict[str, list[Teacher | PlannedRow]],
) -> list[ClassUnit]:
    existing = list(
        db.scalars(select(ClassUnit).where(ClassUnit.semester_id == semester_id)).all()
    )
    by_name: dict[str, list[ClassUnit]] = defaultdict(list)
    for class_unit in existing:
        by_name[class_unit.name].append(class_unit)
    for row in rows:
        if row.errors:
            continue
        homeroom_name = row.values["homeroom_name"]
        if homeroom_name:
            teacher_matches = teacher_candidates.get(homeroom_name, [])
            if not teacher_matches:
                row.add_error("班主任", f"教师「{homeroom_name}」不存在")
            elif len(teacher_matches) > 1:
                row.add_error("班主任", f"教师「{homeroom_name}」有多位同名记录，无法确定引用")
        if row.errors:
            continue
        class_matches = by_name[row.values["name"]]
        if len(class_matches) > 1:
            row.add_error("班名", "现有数据中有多个同名班级，无法确定更新对象")
            continue
        if not class_matches:
            row.status = "new"
            continue
        class_unit = class_matches[0]
        row.existing = class_unit
        changes: list[dict[str, Any]] = []
        _add_change(changes, "年级", class_unit.grade, row.values["grade"])
        _add_change(
            changes,
            "学制",
            TRACK_LABEL_BY_TYPE.get(class_unit.track, class_unit.track),
            TRACK_LABEL_BY_TYPE.get(row.values["track"], row.values["track"]),
        )
        _add_change(changes, "专业/班级类别", class_unit.department, row.values["department"])
        before_homeroom = class_unit.homeroom_teacher.name if class_unit.homeroom_teacher else None
        _add_change(changes, "班主任", before_homeroom, homeroom_name)
        _add_change(changes, "人数", class_unit.student_count, row.values["student_count"])
        row.changes = changes
        row.status = "changed" if changes else "unchanged"
    return existing


def _classify_rooms(
    db: Session,
    semester_id: int,
    rows: list[PlannedRow],
    subject_candidates: dict[str, list[Subject | PlannedRow]],
) -> list[Room]:
    existing = list(db.scalars(select(Room).where(Room.semester_id == semester_id)).all())
    by_name: dict[str, list[Room]] = defaultdict(list)
    for room in existing:
        by_name[room.name].append(room)
    for row in rows:
        if row.errors:
            continue
        _validate_subject_references(
            row, "适用科目", row.values["subject_names"], subject_candidates
        )
        if row.errors:
            continue
        matches = by_name[row.values["name"]]
        if len(matches) > 1:
            row.add_error("名称", "现有数据中有多个同名教室/场地，无法确定更新对象")
            continue
        if not matches:
            row.status = "new"
            continue
        room = matches[0]
        row.existing = room
        changes: list[dict[str, Any]] = []
        _add_change(
            changes,
            "类型",
            ROOM_LABEL_BY_TYPE.get(room.room_type, room.room_type),
            ROOM_LABEL_BY_TYPE.get(row.values["room_type"], row.values["room_type"]),
        )
        _add_change(changes, "容量", room.capacity, row.values["capacity"])
        _add_change(
            changes,
            "适用科目",
            sorted(subject.name for subject in room.subjects),
            sorted(row.values["subject_names"]),
        )
        row.changes = changes
        row.status = "changed" if changes else "unchanged"
    return existing


def _database_snapshot(db: Session, semester_id: int) -> dict[str, Any]:
    subjects = list(
        db.scalars(select(Subject).where(Subject.semester_id == semester_id)).all()
    )
    teachers = list(
        db.scalars(select(Teacher).where(Teacher.semester_id == semester_id)).all()
    )
    classes = list(
        db.scalars(select(ClassUnit).where(ClassUnit.semester_id == semester_id)).all()
    )
    rooms = list(db.scalars(select(Room).where(Room.semester_id == semester_id)).all())
    return {
        "subjects": [
            [
                item.id,
                item.name,
                item.domain,
                item.required_room_type,
                item.default_block_size,
                item.is_major,
            ]
            for item in sorted(subjects, key=lambda value: value.id)
        ],
        "teachers": [
            [
                item.id,
                item.name,
                item.id_last4,
                item.base_periods,
                item.admin_title,
                item.admin_reduction,
                item.is_external,
                item.email,
                item.phone,
                item.line_id,
                sorted(subject.id for subject in item.subjects),
            ]
            for item in sorted(teachers, key=lambda value: value.id)
        ],
        "classes": [
            [
                item.id,
                item.grade,
                item.name,
                item.track,
                item.department,
                item.student_count,
                item.homeroom_teacher_id,
            ]
            for item in sorted(classes, key=lambda value: value.id)
        ],
        "rooms": [
            [
                item.id,
                item.name,
                item.room_type,
                item.capacity,
                sorted(subject.id for subject in item.subjects),
            ]
            for item in sorted(rooms, key=lambda value: value.id)
        ],
    }


def _fingerprint(file_bytes: bytes, database_snapshot: dict[str, Any]) -> str:
    digest = hashlib.sha256()
    digest.update(f"combined-import:{PLAN_VERSION}:".encode())
    digest.update(hashlib.sha256(file_bytes).digest())
    digest.update(json.dumps(database_snapshot, sort_keys=True, separators=(",", ":")).encode())
    return digest.hexdigest()


def build_plan(db: Session, semester_id: int, file_bytes: bytes) -> ImportPlan:
    source, structural_rows = _workbook_rows(file_bytes)
    rows = {
        "subjects": _parse_subjects(source["subjects"]),
        "teachers": _parse_teachers(source["teachers"]),
        "classes": _parse_classes(source["classes"]),
        "rooms": _parse_rooms(source["rooms"]),
    }
    for structural in structural_rows:
        rows[structural.entity].insert(0, structural)

    snapshot = _database_snapshot(db, semester_id)
    existing_subjects = _classify_subjects(db, semester_id, rows["subjects"])
    subject_candidates = _subject_candidates(existing_subjects, rows["subjects"])
    existing_teachers = _classify_teachers(
        db, semester_id, rows["teachers"], subject_candidates
    )
    teacher_candidates = _teacher_candidates(existing_teachers, rows["teachers"])
    _classify_classes(db, semester_id, rows["classes"], teacher_candidates)
    _classify_rooms(db, semester_id, rows["rooms"], subject_candidates)
    return ImportPlan(semester_id, _fingerprint(file_bytes, snapshot), rows)


def _subject_map(db: Session, semester_id: int) -> dict[str, Subject]:
    subjects = db.scalars(select(Subject).where(Subject.semester_id == semester_id)).all()
    return {subject.name: subject for subject in subjects}


def _teacher_map(db: Session, semester_id: int) -> dict[str, Teacher]:
    teachers = db.scalars(select(Teacher).where(Teacher.semester_id == semester_id)).all()
    result: dict[str, Teacher] = {}
    for teacher in teachers:
        if teacher.name not in result:
            result[teacher.name] = teacher
    return result


def apply_plan(db: Session, plan: ImportPlan) -> dict[str, Any]:
    if plan.errors:
        raise ValueError("导入计划仍有冲突")
    created = {entity: 0 for entity in SHEET_DEFS}
    updated = {entity: 0 for entity in SHEET_DEFS}
    unchanged = {entity: 0 for entity in SHEET_DEFS}

    for row in plan.rows["subjects"]:
        values = row.values
        if row.status == "new":
            subject = Subject(semester_id=plan.semester_id, **values)
            db.add(subject)
            row.applied = subject
            created["subjects"] += 1
        else:
            existing_subject = row.existing
            assert isinstance(existing_subject, Subject)
            row.applied = existing_subject
            if row.status == "changed":
                for key, value in values.items():
                    setattr(existing_subject, key, value)
                updated["subjects"] += 1
            else:
                unchanged["subjects"] += 1
    db.flush()
    subjects = _subject_map(db, plan.semester_id)

    for row in plan.rows["teachers"]:
        values = row.values
        scalar_values = {
            key: value for key, value in values.items() if key != "subject_names"
        }
        subject_objects = [subjects[name] for name in values["subject_names"]]
        if row.status == "new":
            teacher = Teacher(
                semester_id=plan.semester_id,
                **scalar_values,
                subjects=subject_objects,
            )
            db.add(teacher)
            row.applied = teacher
            created["teachers"] += 1
        else:
            existing_teacher = row.existing
            assert isinstance(existing_teacher, Teacher)
            row.applied = existing_teacher
            if row.status == "changed":
                for key, value in scalar_values.items():
                    setattr(existing_teacher, key, value)
                existing_teacher.subjects = subject_objects
                updated["teachers"] += 1
            else:
                unchanged["teachers"] += 1
    db.flush()
    teachers = _teacher_map(db, plan.semester_id)

    for row in plan.rows["classes"]:
        values = row.values
        homeroom_name = values["homeroom_name"]
        scalar_values = {
            key: value for key, value in values.items() if key != "homeroom_name"
        }
        homeroom_id = teachers[homeroom_name].id if homeroom_name else None
        if row.status == "new":
            class_unit = ClassUnit(
                semester_id=plan.semester_id,
                **scalar_values,
                homeroom_teacher_id=homeroom_id,
            )
            db.add(class_unit)
            row.applied = class_unit
            created["classes"] += 1
        else:
            existing_class = row.existing
            assert isinstance(existing_class, ClassUnit)
            row.applied = existing_class
            if row.status == "changed":
                for key, value in scalar_values.items():
                    setattr(existing_class, key, value)
                existing_class.homeroom_teacher_id = homeroom_id
                updated["classes"] += 1
            else:
                unchanged["classes"] += 1

    for row in plan.rows["rooms"]:
        values = row.values
        scalar_values = {
            key: value for key, value in values.items() if key != "subject_names"
        }
        subject_objects = [subjects[name] for name in values["subject_names"]]
        if row.status == "new":
            room = Room(
                semester_id=plan.semester_id,
                **scalar_values,
                subjects=subject_objects,
            )
            db.add(room)
            row.applied = room
            created["rooms"] += 1
        else:
            existing_room = row.existing
            assert isinstance(existing_room, Room)
            row.applied = existing_room
            if row.status == "changed":
                for key, value in scalar_values.items():
                    setattr(existing_room, key, value)
                existing_room.subjects = subject_objects
                updated["rooms"] += 1
            else:
                unchanged["rooms"] += 1

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "total_created": sum(created.values()),
        "total_updated": sum(updated.values()),
        "total_unchanged": sum(unchanged.values()),
    }
