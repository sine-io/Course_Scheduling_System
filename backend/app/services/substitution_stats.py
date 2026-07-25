"""代课课时月结统计(M4-5,architecture.md §5.4)。

回答「这个月每位老师代了几节、其中几节要计课时费」。真相仍是 `substitution` 列
(处理方式决定)+ `affected_period`(受影响节次快照),这里依教师汇总。

**两个数字**:
- 代课节数:该教师接手的所有处理方式(代课/调课/合班),即他实际处理了几节。
- 计费节数:其中 `counts_toward_hours` 为真者。合班/自习默认不计、代课默认计(可覆盖)。

**跨月假单自动拆月**:以每一个 `affected_period` 自己的日期分月,不是以假单分月。
王师请 1/30~2/2 的假,1 月的节次进 1 月报表、2 月的进 2 月——不必特别处理。

**销假的节次不计**:销假会把未完成的节次转为 `cancelled`(那堂课没上),故排除;
已完成(completed)的节次即使事后销假仍保留(课上过了,课时照算)。
"""

import io
from dataclasses import dataclass, field
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.leave import AffectedPeriod, AffectedStatus
from app.models.substitution import Substitution
from app.services import school_rules

_Date = date


@dataclass(frozen=True, slots=True)
class StatDetail:
    """一列代课明细:某教师某节的处理方式。"""

    handler_teacher_id: int
    handler_name: str
    date: _Date
    period_no: int
    period_name: str
    class_names: str
    subject_name: str
    absent_teacher_name: str
    leave_type: str
    leave_type_label: str
    sub_type: str
    sub_type_label: str
    counts_toward_hours: bool
    funding_source: str


@dataclass
class TeacherSummary:
    teacher_id: int
    teacher_name: str
    handled_count: int = 0  # 代课节数(所有接手处理方式)
    billable_count: int = 0  # 计费节数(counts_toward_hours 为真)


@dataclass
class MonthlyReport:
    year: int
    month: int
    summaries: list[TeacherSummary] = field(default_factory=list)
    details: list[StatDetail] = field(default_factory=list)


def _next_month(year: int, month: int) -> date:
    return date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)


def monthly_report(
    db: Session,
    semester_id: int,
    year: int,
    month: int,
    *,
    teacher_id: int | None = None,
) -> MonthlyReport:
    """某月的代课课时统计。指定 teacher_id 则只统计该教师(教师个人查询用)。"""
    month_start = date(year, month, 1)
    month_end = _next_month(year, month)

    stmt = (
        select(Substitution, AffectedPeriod)
        .join(AffectedPeriod, Substitution.affected_period_id == AffectedPeriod.id)
        .where(
            Substitution.semester_id == semester_id,
            Substitution.handler_teacher_id.isnot(None),
            AffectedPeriod.status != AffectedStatus.cancelled.value,
            AffectedPeriod.date >= month_start,
            AffectedPeriod.date < month_end,
        )
    )
    if teacher_id is not None:
        stmt = stmt.where(Substitution.handler_teacher_id == teacher_id)

    report = MonthlyReport(year=year, month=month)
    summaries: dict[int, TeacherSummary] = {}

    for sub, ap in db.execute(stmt).all():
        handler = sub.handler
        if handler is None:  # handler 已被移除(SET NULL 尚未反映在关联)
            continue
        leave = ap.leave_request
        report.details.append(
            StatDetail(
                handler_teacher_id=handler.id,
                handler_name=handler.name,
                date=ap.date,
                period_no=ap.period_no,
                period_name=ap.period_name,
                class_names=ap.class_names,
                subject_name=ap.subject_name,
                absent_teacher_name=leave.teacher.name if leave.teacher else "(已移除)",
                leave_type=leave.leave_type,
                leave_type_label=school_rules.leave_type_label(leave.leave_type),
                sub_type=sub.type,
                sub_type_label=school_rules.substitution_type_label(sub.type),
                counts_toward_hours=sub.counts_toward_hours,
                funding_source=sub.funding_source,
            )
        )
        s = summaries.get(handler.id)
        if s is None:
            s = TeacherSummary(teacher_id=handler.id, teacher_name=handler.name)
            summaries[handler.id] = s
        s.handled_count += 1
        if sub.counts_toward_hours:
            s.billable_count += 1

    report.details.sort(key=lambda d: (d.handler_name, d.date, d.period_no))
    report.summaries = sorted(summaries.values(), key=lambda s: s.teacher_name)
    return report


def _detail_headers() -> tuple[str, ...]:
    label = school_rules.export_label
    return (
        label("teacher"),
        label("date"),
        label("period"),
        label("class"),
        label("subject"),
        label("absent_teacher"),
        label("leave_type"),
        label("disposition"),
        label("billable"),
        label("funding_source"),
    )


def _summary_headers() -> tuple[str, ...]:
    label = school_rules.export_label
    return (label("teacher"), label("substitution_periods"), label("billable_periods"))


def build_workbook(report: MonthlyReport) -> bytes:
    """导出两张表:汇总(每位教师)+ 明细(逐节)。返回 xlsx bytes。"""
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = school_rules.export_label("summary")
    ws_sum.append(list(_summary_headers()))
    for s in report.summaries:
        ws_sum.append([s.teacher_name, s.handled_count, s.billable_count])

    ws_detail = wb.create_sheet(school_rules.export_label("detail"))
    ws_detail.append(list(_detail_headers()))
    for d in report.details:
        ws_detail.append(
            [
                d.handler_name,
                d.date.isoformat(),
                d.period_name,
                d.class_names,
                d.subject_name,
                d.absent_teacher_name,
                d.leave_type_label,
                d.sub_type_label,
                (
                    school_rules.export_label("yes")
                    if d.counts_toward_hours
                    else school_rules.export_label("no")
                ),
                d.funding_source,
            ]
        )

    for ws in (ws_sum, ws_detail):
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
