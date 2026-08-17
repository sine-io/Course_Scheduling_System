"""组合基础数据导入的公开 API 契约测试。"""

import io

import pytest
from openpyxl import Workbook, load_workbook

from app.api.imports import XLSX_MIME
from app.models.basedata import ClassUnit, Room, Subject, Teacher
from app.models.user import Role, User
from tests.conftest import make_user

PW = "password123"

SHEETS = {
    "科目": ["名称", "领域/类别", "所需教室/场地类型", "默认连堂", "主科"],
    "教师": [
        "姓名",
        "身份后四位",
        "任教科目",
        "基本课时",
        "行政职务",
        "行政减课",
        "外聘",
        "邮箱",
        "手机号",
        "其他联系方式",
    ],
    "班级": ["年级", "班名", "学制", "专业/班级类别", "班主任", "人数"],
    "教室": ["名称", "类型", "容量", "适用科目"],
}


@pytest.fixture
def scheduler_env(env):
    client, db = env
    make_user(db, "setup", PW, roles=[Role.scheduler])
    client.post("/api/auth/login", json={"username": "setup", "password": PW})
    response = client.post(
        "/api/semesters",
        json={
            "academic_year": 2026,
            "term": 1,
            "start_date": "2026-09-01",
            "end_date": "2027-01-20",
        },
    )
    assert response.status_code == 201, response.json()
    semester = response.json()
    return client, db, semester["id"]


def combined_workbook(**rows_by_sheet: list[list[object]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, columns in SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(columns)
        sheet.append(["说明"] * len(columns))
        sheet.append(["示例"] * len(columns))
        for row in rows_by_sheet.get(sheet_name, []):
            sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def post_workbook(client, path: str, semester_id: int, content: bytes, data=None):
    return client.post(
        f"/api/import/setup/{path}?semester_id={semester_id}",
        data=data,
        files={"file": ("school-setup.xlsx", content, XLSX_MIME)},
    )


def test_combined_template_exposes_four_sheets_without_teacher_accounts(scheduler_env):
    client, _, _ = scheduler_env

    response = client.get("/api/import/setup/template")

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == list(SHEETS)
    for sheet_name, expected_columns in SHEETS.items():
        values = [cell.value for cell in workbook[sheet_name][1]]
        assert values == expected_columns
    assert "登录账号" not in {
        cell.value for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row
    }


def test_preview_resolves_cross_sheet_references_and_writes_nothing(scheduler_env):
    client, db, semester_id = scheduler_env
    user_count = db.query(User).count()
    content = combined_workbook(
        **{
            "科目": [["数学", "数学", "普通教室", 2, "是"]],
            "教师": [["王老师", "1234", "数学", 18, "", 0, "否", "", "", ""]],
            "班级": [[7, "七年级1班", "初中", "", "王老师", 42]],
            "教室": [["实验室A", "专用教室", 48, "数学"]],
        }
    )

    response = post_workbook(client, "preview", semester_id, content)

    assert response.status_code == 200
    preview = response.json()
    assert preview["can_commit"] is True
    assert preview["counts"] == {"new": 4, "unchanged": 0, "changed": 0, "conflict": 0}
    assert db.query(Subject).count() == 0
    assert db.query(Teacher).count() == 0
    assert db.query(ClassUnit).count() == 0
    assert db.query(Room).count() == 0

    committed = post_workbook(
        client,
        "commit",
        semester_id,
        content,
        data={"fingerprint": preview["fingerprint"], "confirm_changes": "false"},
    )

    assert committed.status_code == 200
    assert committed.json()["created"] == {
        "subjects": 1,
        "teachers": 1,
        "classes": 1,
        "rooms": 1,
    }
    teacher = db.query(Teacher).one()
    class_unit = db.query(ClassUnit).one()
    room = db.query(Room).one()
    assert [subject.name for subject in teacher.subjects] == ["数学"]
    assert class_unit.homeroom_teacher_id == teacher.id
    assert [subject.name for subject in room.subjects] == ["数学"]
    assert db.query(User).count() == user_count


def test_changed_rows_require_confirmation_and_then_update(scheduler_env):
    client, db, semester_id = scheduler_env
    subject = Subject(
        semester_id=semester_id,
        name="数学",
        domain="数学",
        required_room_type="normal",
        default_block_size=1,
        is_major=True,
    )
    teacher = Teacher(
        semester_id=semester_id,
        name="王老师",
        id_last4="1234",
        base_periods=16,
        subjects=[subject],
    )
    db.add_all([subject, teacher])
    db.commit()
    content = combined_workbook(
        **{
            "科目": [["数学", "数学", "普通教室", 1, "是"]],
            "教师": [["王老师", "1234", "数学", 18, "", 0, "否", "", "", ""]],
        }
    )

    preview = post_workbook(client, "preview", semester_id, content).json()

    assert preview["counts"] == {"new": 0, "unchanged": 1, "changed": 1, "conflict": 0}
    changed = next(
        row
        for sheet in preview["sheets"]
        for row in sheet["rows"]
        if row["status"] == "changed"
    )
    assert changed["identity"] == "王老师（1234）"
    assert changed["changes"] == [{"field": "基本课时", "before": 16, "after": 18}]

    rejected = post_workbook(
        client,
        "commit",
        semester_id,
        content,
        data={"fingerprint": preview["fingerprint"], "confirm_changes": "false"},
    )
    assert rejected.status_code == 409
    assert rejected.json()["detail"]["code"] == "combined_import_changes_unconfirmed"
    db.refresh(teacher)
    assert teacher.base_periods == 16

    committed = post_workbook(
        client,
        "commit",
        semester_id,
        content,
        data={"fingerprint": preview["fingerprint"], "confirm_changes": "true"},
    )
    assert committed.status_code == 200
    assert committed.json()["updated"]["teachers"] == 1
    db.refresh(teacher)
    assert teacher.base_periods == 18


def test_ambiguous_teacher_identity_is_a_located_conflict(scheduler_env):
    client, db, semester_id = scheduler_env
    db.add_all(
        [
            Teacher(semester_id=semester_id, name="王老师", id_last4="1234"),
            Teacher(semester_id=semester_id, name="王老师", id_last4="5678"),
        ]
    )
    db.commit()
    content = combined_workbook(
        **{"教师": [["王老师", "", "", 18, "", 0, "否", "", "", ""]]}
    )

    preview = post_workbook(client, "preview", semester_id, content).json()

    assert preview["can_commit"] is False
    assert preview["counts"]["conflict"] == 1
    assert preview["errors"] == [
        {
            "sheet": "教师",
            "row": 4,
            "field": "身份后四位",
            "message": "现有数据中有多位同名教师，请填写身份后四位以明确对应关系",
        }
    ]


def test_conflict_blocks_the_whole_commit(scheduler_env):
    client, db, semester_id = scheduler_env
    content = combined_workbook(
        **{
            "科目": [["数学", "数学", "普通教室", 1, "是"]],
            "教师": [["王老师", "1234", "不存在", 18, "", 0, "否", "", "", ""]],
        }
    )
    preview = post_workbook(client, "preview", semester_id, content).json()
    assert preview["can_commit"] is False

    response = post_workbook(
        client,
        "commit",
        semester_id,
        content,
        data={"fingerprint": preview["fingerprint"], "confirm_changes": "true"},
    )

    assert response.status_code == 409
    assert response.json()["detail"]["code"] == "combined_import_conflicts"
    assert db.query(Subject).count() == 0
    assert db.query(Teacher).count() == 0


def test_commit_rejects_a_stale_preview(scheduler_env):
    client, _, semester_id = scheduler_env
    content = combined_workbook(**{"科目": [["数学", "数学", "", 1, "是"]]})
    preview = post_workbook(client, "preview", semester_id, content).json()
    created = client.post(
        f"/api/subjects?semester_id={semester_id}",
        json={"name": "语文", "is_major": True},
    )
    assert created.status_code == 201

    response = post_workbook(
        client,
        "commit",
        semester_id,
        content,
        data={"fingerprint": preview["fingerprint"], "confirm_changes": "false"},
    )

    assert response.status_code == 409
    assert response.json()["detail"]["code"] == "combined_import_preview_stale"
