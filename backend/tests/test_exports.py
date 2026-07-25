"""M5-1:课表导出。Excel 以读回比对验证内容;HTML(PDF 来源)验中文与版面;
全校总表/批量/RBAC。PDF/PNG 的 worker 渲染于真实环境另行验证。
"""

import io
import zipfile

import pytest
from openpyxl import load_workbook

from app.models.user import Role
from app.services import timetable_export as tex
from tests.api_helpers import create_api_semester
from tests.conftest import make_user
from tests.dates import SEM_END, SEM_START  # 日期统一由执行当日推算,不硬编
from tests.test_substitutions import _World

PW = "password123"


@pytest.fixture
def w(env):
    client, db = env
    make_user(db, "s", PW, roles=[Role.scheduler])
    client.post("/api/auth/login", json={"username": "s", "password": PW})
    sid = create_api_semester(
        client,
        ready=True,
        start_date=SEM_START.isoformat(),
        end_date=SEM_END.isoformat(),
    )["id"]
    world = _World(client, db, sid)
    world.teacher("王师", ["语文"])
    world.place("王师", "语文", "701", 0)   # 周三第一节
    world.publish()
    return world


def _cells(ws) -> set[str]:
    out: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v:
                out.add(str(v))
    return out


def _xlsx(resp):
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    return load_workbook(io.BytesIO(resp.content))


# ── 内容一致(Excel 读回)─────────────────────────────────────
def test_class_excel_matches_grid(w):
    cid = w.classes["701"]
    r = w.client.get(f"/api/export/timetable{w.q}&view=class&target_id={cid}&fmt=xlsx")
    wb = _xlsx(r)
    ws = wb.active
    cells = _cells(ws)
    assert "701 课表" in " ".join(cells) or any("课表" in c for c in cells)
    # 周三第一节有「语文」与「王师」
    joined = "\n".join(cells)
    assert "语文" in joined
    assert "王师" in joined
    assert "星期三" in cells  # 表头星期
    assert "第一节" in cells  # 节次名


def test_teacher_excel_shows_class_not_teacher(w):
    tid = w.teachers["王师"]
    r = w.client.get(f"/api/export/timetable{w.q}&view=teacher&target_id={tid}&fmt=xlsx")
    wb = _xlsx(r)
    joined = "\n".join(_cells(wb.active))
    assert "语文" in joined
    assert "701" in joined       # 教师视角格内显示班级


def test_room_export_ok_even_without_room(w):
    # 建一个没有课的教室/场地也应导出成功(空课表)
    rid = w.client.post(f"/api/rooms{w.q}", json={"name": "101教室"}).json()["id"]
    r = w.client.get(f"/api/export/timetable{w.q}&view=room&target_id={rid}&fmt=xlsx")
    assert r.status_code == 200


# ── HTML(PDF 来源)─────────────────────────────────────────
def test_grid_html_has_chinese_and_layout(w):
    cid = w.classes["701"]
    grid, meta = tex.build_grid(w.db, w.sid, "class", cid)
    html = tex.grid_to_html(grid, meta)
    assert "语文" in html and "王师" in html
    assert "星期三" in html and "第一节" in html
    assert "A4 portrait" in html
    assert meta.school_name in html


# ── 全校总表 / 批量 ─────────────────────────────────────────
def test_school_workbook_one_sheet_per_class(w):
    # 学期已有 701 与占位班 900 → 全校总表中每班一个工作表
    r = w.client.get(f"/api/export/school.xlsx{w.q}")
    wb = _xlsx(r)
    assert len(wb.sheetnames) >= 2
    assert any("701" in s for s in wb.sheetnames)


def test_batch_zip_has_per_class_files(w):
    r = w.client.get(f"/api/export/batch.zip{w.q}")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/zip"
    z = zipfile.ZipFile(io.BytesIO(r.content))
    names = z.namelist()
    assert any(n.endswith(".xlsx") for n in names)
    assert any("701" in n for n in names)


# ── 边界 / RBAC ─────────────────────────────────────────────
def test_export_unpublished_semester_404(env):
    client, db = env
    make_user(db, "s2", PW, roles=[Role.scheduler])
    client.post("/api/auth/login", json={"username": "s2", "password": PW})
    sid = create_api_semester(
        client,
        academic_year=2027,
        start_date=SEM_START.isoformat(),
        end_date=SEM_END.isoformat(),
    )["id"]
    cid = client.post(f"/api/class-units?semester_id={sid}",
                      json={"grade": 7, "name": "701", "track": "junior_high"}).json()["id"]
    r = client.get(f"/api/export/timetable?semester_id={sid}&view=class&target_id={cid}&fmt=xlsx")
    assert r.status_code == 404


def test_batch_60_classes_under_60s(db):
    """验收②:60 班批量导出 < 60 秒(Excel 生成为 CPU-bound,与数据库无关)。"""
    import time

    from app.models.timetable import Timetable, TimetableStatus
    from tests.fixtures import build_large_school

    fx = build_large_school(db, num_classes=60)
    db.add(Timetable(
        semester_id=fx.semester_id, name="正式课表",
        status=TimetableStatus.published.value))
    db.commit()

    t0 = time.perf_counter()
    data = tex.class_batch_zip(db, fx.semester_id)
    elapsed = time.perf_counter() - t0
    assert elapsed < 60, f"批量导出耗时 {elapsed:.1f}s"
    z = zipfile.ZipFile(io.BytesIO(data))
    assert len(z.namelist()) == 60


def test_teacher_can_export_single_but_not_batch(w):
    make_user(w.db, "t", PW, roles=[Role.teacher])
    w.client.post("/api/auth/logout")
    w.client.post("/api/auth/login", json={"username": "t", "password": PW})
    cid = w.classes["701"]
    single = w.client.get(f"/api/export/timetable{w.q}&view=class&target_id={cid}&fmt=xlsx")
    assert single.status_code == 200          # 单一课表全校可查可导出
    assert w.client.get(f"/api/export/school.xlsx{w.q}").status_code == 403
    assert w.client.get(f"/api/export/batch.zip{w.q}").status_code == 403
