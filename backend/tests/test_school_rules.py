"""统一学校规则、应用配置、校历和导出文案回归测试。"""

import io
from datetime import date

from openpyxl import load_workbook

from app.models.calendar import CalendarExceptionKind, SemesterCalendarException
from app.models.semester import Semester
from app.models.user import Role
from app.services import calendar as calendar_service
from app.services.school_rules import (
    ROLE_DISPLAY_NAMES,
    TIMEZONE,
    affected_status_label,
    format_semester_label,
    leave_type_label,
    substitution_type_label,
    validate_academic_year,
)
from app.services.timetable_export import Grid, Meta, grids_to_xlsx
from tests.conftest import make_user

PW = "password123"


def test_school_rules_use_simplified_chinese_and_gregorian_years():
    assert TIMEZONE == "Asia/Shanghai"
    assert ROLE_DISPLAY_NAMES["scheduler"] == "排课管理员"
    assert format_semester_label(2026, 1) == "2026-2027学年第一学期"
    assert format_semester_label(2026, 2) == "2026-2027学年第二学期"
    assert leave_type_label("bereavement") == "丧假"
    assert affected_status_label("pending") == "待处理"
    assert substitution_type_label("merge") == "合班"


def test_academic_year_range_is_1900_to_2100():
    validate_academic_year(1900)
    validate_academic_year(2100)
    for invalid in (1899, 2101):
        try:
            validate_academic_year(invalid)
        except ValueError as exc:
            assert "1900 至 2100" in str(exc)
        else:
            raise AssertionError("超出范围的学年必须被拒绝")


def test_calendar_effective_weekday_handles_closure_and_makeup_day(db):
    semester = Semester(
        academic_year=2026,
        term=1,
        start_date=date(2026, 9, 1),
        end_date=date(2027, 1, 31),
    )
    db.add(semester)
    db.flush()
    db.add_all(
        [
            SemesterCalendarException(
                semester_id=semester.id,
                date=date(2026, 9, 7),
                kind=CalendarExceptionKind.no_instruction.value,
            ),
            SemesterCalendarException(
                semester_id=semester.id,
                date=date(2026, 9, 12),
                kind=CalendarExceptionKind.makeup_instruction.value,
                makeup_weekday=1,
            ),
        ]
    )
    db.flush()
    assert calendar_service.effective_weekday(db, semester.id, date(2026, 9, 7)) is None
    assert calendar_service.effective_weekday(db, semester.id, date(2026, 9, 12)) == 1
    assert calendar_service.effective_weekday(db, semester.id, date(2026, 9, 8)) == 2


def test_public_app_config_has_only_supported_fields(env):
    client, _ = env

    response = client.get("/api/app-config")

    assert response.status_code == 200
    body = response.json()
    assert set(body) == {
        "school_name",
        "timezone",
        "role_display_names",
        "academic_year",
    }
    assert body["timezone"] == "Asia/Shanghai"
    assert body["role_display_names"]["scheduler"] == "排课管理员"
    assert body["academic_year"] == {
        "storage": "start_year",
        "min": 1900,
        "max": 2100,
        "label_format": "{year}-{next_year}学年{term_label}",
        "term_labels": {"1": "第一学期", "2": "第二学期"},
    }


def test_calendar_api_crud_and_readiness_confirmation(env):
    client, db = env
    make_user(db, "scheduler", PW, roles=[Role.scheduler])
    login = client.post("/api/auth/login", json={"username": "scheduler", "password": PW})
    assert login.status_code == 200

    semester_response = client.post(
        "/api/semesters",
        json={
            "academic_year": 2026,
            "term": 1,
            "template_key": "junior_high_draft",
            "start_date": "2026-09-01",
            "end_date": "2027-01-31",
        },
    )
    assert semester_response.status_code == 201
    semester = semester_response.json()
    sid = semester["id"]

    invalid = client.post(
        f"/api/semesters/{sid}/calendar-exceptions",
        json={"date": "2026-09-12", "kind": "makeup_instruction"},
    )
    assert invalid.status_code in {400, 422}

    created = client.post(
        f"/api/semesters/{sid}/calendar-exceptions",
        json={"date": "2026-09-12", "kind": "makeup_instruction", "makeup_weekday": 1},
    )
    assert created.status_code == 201
    exception_id = created.json()["id"]

    blocked = client.post(f"/api/semesters/{sid}/readiness")
    assert blocked.status_code == 409
    assert blocked.json()["detail"]["code"] == "semester_not_ready"

    table_id = semester["period_tables"][0]["id"]
    periods = [
        {"weekday": weekday, "period_no": 1, "name": "第一节", "type": "regular"}
        for weekday in range(1, 6)
    ]
    assert client.put(f"/api/period-tables/{table_id}/periods", json=periods).status_code == 200
    ready = client.post(f"/api/semesters/{sid}/readiness")
    assert ready.status_code == 200
    assert ready.json()["ready"] is True

    updated = client.patch(
        f"/api/calendar-exceptions/{exception_id}",
        json={"kind": "no_instruction", "makeup_weekday": None, "note": "运动会"},
    )
    assert updated.status_code == 200
    assert updated.json()["kind"] == "no_instruction"
    assert client.get(f"/api/semesters/{sid}/readiness").json()["ready"] is False
    assert client.delete(f"/api/calendar-exceptions/{exception_id}").status_code == 204


def test_export_metadata_uses_simplified_chinese_labels():
    grid = Grid(title="七年级一班课表", num_weekdays=5)
    data = grids_to_xlsx(
        [grid],
        Meta("示范学校", "2026-2027学年第一学期", "草稿", date(2026, 9, 1)),
    )
    sheet = load_workbook(io.BytesIO(data)).active
    assert sheet["A2"].value.endswith("打印日期：2026-09-01")
    assert [sheet.cell(3, col).value for col in range(1, 4)] == [
        "节次",
        "星期一",
        "星期二",
    ]
