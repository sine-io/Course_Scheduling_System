"""CN deployment profile and calendar foundation regression tests."""

import io
from datetime import date

from openpyxl import load_workbook

from app.core.config import Settings
from app.models.app_setting import AppSetting
from app.models.calendar import CalendarExceptionKind, SemesterCalendarException
from app.models.semester import Semester
from app.models.user import Role
from app.services import calendar as calendar_service
from app.services.deployment_profile import ProfileMismatchError, ensure_locked_profile
from app.services.localization import (
    format_semester_label,
    leave_type_label,
    localize_payload,
    localize_text,
    public_profile,
    substitution_type_label,
)
from app.services.timetable_export import Grid, Meta, grids_to_xlsx
from tests.conftest import make_user

PW = "password123"


def test_mainland_profile_formats_academic_year_and_public_config():
    assert format_semester_label(2026, 1, "cn_mainland") == "2026-2027学年第一学期"
    config = public_profile("cn_mainland")
    assert config["locale"] == "zh-CN"
    assert config["timezone"] == "Asia/Shanghai"
    assert config["academic_year"]["min"] == 1900
    assert config["role_display_names"]["scheduler"] == "教务员"
    assert leave_type_label("bereavement", "cn_mainland") == "丧假"
    assert substitution_type_label("merge", "cn_mainland") == "合班"


def test_mainland_normalizes_legacy_user_facing_messages():
    assert localize_text("帳號或密碼錯誤", "cn_mainland") == "账号或密码错误"
    assert (
        localize_text("跑班群組每週節數不一致", "cn_mainland")
        == "走班群组每周节数不一致"
    )
    payload = {
        "code": "invalid_assignment",
        "message": "請先建立導師與專科教室",
        "issues": [{"message": "資料未通過排課前置檢查"}],
    }
    assert localize_payload(payload, "cn_mainland") == {
        "code": "invalid_assignment",
        "message": "请先建立班主任与专用教室",
        "issues": [{"message": "资料未通过排课前置检查"}],
    }


def test_settings_reject_unknown_profile():
    try:
        Settings(_env_file=None, secret_key="real", school_profile="unknown")
    except ValueError as exc:
        assert "SCHOOL_PROFILE" in str(exc)
    else:
        raise AssertionError("unknown profile must be rejected")


def test_profile_is_locked_and_cannot_cross_region(db, monkeypatch):
    monkeypatch.setattr("app.services.deployment_profile.settings.school_profile", "cn_mainland")
    assert ensure_locked_profile(db) == "cn_mainland"
    db.commit()
    monkeypatch.setattr("app.services.deployment_profile.settings.school_profile", "tw_k12")
    try:
        ensure_locked_profile(db)
    except ProfileMismatchError as exc:
        assert exc.locked == "cn_mainland"
    else:
        raise AssertionError("profile switch must be rejected")


def test_calendar_effective_weekday_handles_stop_and_makeup(db):
    semester = Semester(
        academic_year=2026,
        term=1,
        start_date=date(2026, 9, 1),
        end_date=date(2027, 1, 31),
    )
    db.add(semester)
    db.flush()
    db.add_all([
        SemesterCalendarException(
            semester_id=semester.id,
            date=date(2026, 9, 7),
            kind=CalendarExceptionKind.no_instruction.value,
        ),
        SemesterCalendarException(
            semester_id=semester.id,
            date=date(2026, 9, 12),
            kind=CalendarExceptionKind.makeup_instruction.value,
            makeup_weekday=1,
        ),
    ])
    db.flush()
    assert calendar_service.effective_weekday(db, semester.id, date(2026, 9, 7)) is None
    assert calendar_service.effective_weekday(db, semester.id, date(2026, 9, 12)) == 1
    assert calendar_service.effective_weekday(db, semester.id, date(2026, 9, 8)) == 2


def test_public_app_config_locks_a_new_mainland_deployment(env, monkeypatch):
    client, db = env
    monkeypatch.setattr("app.services.deployment_profile.settings.school_profile", "cn_mainland")
    monkeypatch.setattr("app.services.localization.settings.school_profile", "cn_mainland")

    response = client.get("/api/app-config")

    assert response.status_code == 200
    body = response.json()
    assert body["profile"] == "cn_mainland"
    assert body["locale"] == "zh-CN"
    assert body["academic_year"]["min"] == 1900
    db.expire_all()
    assert db.get(AppSetting, "school_profile").value == "cn_mainland"


def test_calendar_api_crud_and_readiness_confirmation(env, monkeypatch):
    client, db = env
    monkeypatch.setattr("app.services.deployment_profile.settings.school_profile", "cn_mainland")
    monkeypatch.setattr("app.services.localization.settings.school_profile", "cn_mainland")
    make_user(db, "scheduler", PW, roles=[Role.scheduler])
    login = client.post("/api/auth/login", json={"username": "scheduler", "password": PW})
    assert login.status_code == 200

    semester = client.post(
        "/api/semesters",
        json={
            "academic_year": 2026,
            "term": 1,
            "template_key": "cn_junior_high_draft",
            "start_date": "2026-09-01",
            "end_date": "2027-01-31",
        },
    ).json()
    sid = semester["id"]

    invalid = client.post(
        f"/api/semesters/{sid}/calendar-exceptions",
        json={"date": "2026-09-12", "kind": "makeup_instruction"},
    )
    assert invalid.status_code == 422

    created = client.post(
        f"/api/semesters/{sid}/calendar-exceptions",
        json={"date": "2026-09-12", "kind": "makeup_instruction", "makeup_weekday": 1},
    )
    assert created.status_code == 201
    exception_id = created.json()["id"]
    assert created.json()["makeup_weekday"] == 1

    blocked = client.post(f"/api/semesters/{sid}/readiness")
    assert blocked.status_code == 409
    assert blocked.json()["detail"]["code"] == "semester_not_ready"

    table_id = semester["period_tables"][0]["id"]
    assert client.put(
        f"/api/period-tables/{table_id}/periods",
        json=[{"weekday": 1, "period_no": 1, "name": "第一节", "type": "regular"}],
    ).status_code == 200
    ready = client.post(f"/api/semesters/{sid}/readiness")
    assert ready.status_code == 200
    assert ready.json()["ready"] is True

    updated = client.patch(
        f"/api/calendar-exceptions/{exception_id}",
        json={"kind": "no_instruction", "makeup_weekday": None, "note": "运动会"},
    )
    assert updated.status_code == 200
    assert updated.json()["kind"] == "no_instruction"
    assert client.get(f"/api/semesters/{sid}/readiness").json()["ready"] is False
    assert client.delete(f"/api/calendar-exceptions/{exception_id}").status_code == 204


def test_mainland_export_metadata_uses_simplified_labels(monkeypatch):
    monkeypatch.setattr("app.services.localization.settings.school_profile", "cn_mainland")
    grid = Grid(title="701 课表", num_weekdays=5)
    data = grids_to_xlsx(
        [grid],
        Meta("天津示范学校", "2026-2027学年第一学期", "草稿", date(2026, 9, 1)),
    )
    sheet = load_workbook(io.BytesIO(data)).active
    assert sheet["A2"].value.endswith("打印日期:2026-09-01")
    assert [sheet.cell(3, col).value for col in range(1, 4)] == ["节次", "周一", "周二"]
