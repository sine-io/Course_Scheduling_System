"""M4-5:代课课时月结统计。

重点:两个数字(代课节数 vs 计费节数)、不计费项排除(合班/自习)、跨月假单拆月计、
教师个人只看自己、Excel 导出字段。
"""

import io

import pytest
from openpyxl import load_workbook

from app.models.basedata import Teacher
from app.models.user import Role
from tests.api_helpers import create_api_semester
from tests.conftest import make_user

# 日期统一由执行当日推算,不硬编(见 tests/dates.py)。
# CROSS_WED / CROSS_WED2 是相邻但分属不同月份的两个周三,供跨月拆账验证。
from tests.dates import CROSS_WED, CROSS_WED2, SEM_END, SEM_START, WED
from tests.test_substitutions import _World

PW = "password123"


def _other_month(day):
    """一个确定没有任何代课记录的月份(基准日的下个月)。"""
    return (day.year + 1, 1) if day.month == 12 else (day.year, day.month + 1)


@pytest.fixture
def w(env):
    client, db = env
    make_user(db, "s", PW, roles=[Role.scheduler])
    client.post("/api/auth/login", json={"username": "s", "password": PW})
    sid = create_api_semester(
        client,
        ready=True,
        start_date=SEM_START.isoformat(),
        end_date=SEM_END.isoformat(),
    )["id"]
    return _World(client, db, sid)


def _stats(w, year=None, month=None, **params):
    """默认查「基准周那个月」——受影响节次都落在那里。"""
    year = WED.year if year is None else year
    month = WED.month if month is None else month
    qs = "".join(f"&{k}={v}" for k, v in params.items() if v is not None)
    return w.client.get(f"/api/substitution-stats{w.q}&year={year}&month={month}{qs}").json()


# ── 验收①:代课节数 vs 计费节数,不计费项排除 ──────────────
def test_handled_vs_billable_counts(w):
    """陈师代课 1 节(计费)+ 合班 1 节(不计费);自习无处理教师不计入任何人。"""
    w.teacher("王师", ["语文"])
    w.teacher("陈师", ["语文"])
    w.place("王师", "语文", "701", 0)  # 周三第一节
    w.place("王师", "语文", "702", 1)  # 周三第二节
    w.place("王师", "语文", "703", 2)  # 周三第三节
    w.publish()
    ap = w.leave("王师")  # 3 节受影响(同一天)

    w.assign(ap[0]["id"], type="substitute", handler_teacher_id=w.teachers["陈师"])
    w.assign(ap[1]["id"], type="merge", handler_teacher_id=w.teachers["陈师"])
    w.assign(ap[2]["id"], type="self_study")  # 无处理教师

    data = _stats(w)
    chen = next(s for s in data["summaries"] if s["teacher_name"] == "陈师")
    assert chen["handled_count"] == 2   # 代课 + 合班
    assert chen["billable_count"] == 1  # 只有代课计费
    # 自习没有处理教师,不会出现在任何人的统计
    assert all(s["teacher_name"] != "王师" for s in data["summaries"])
    assert len(data["details"]) == 2


def test_detail_columns(w):
    w.teacher("王师", ["语文"])
    w.teacher("陈师", ["语文"])
    w.place("王师", "语文", "701", 0)
    w.publish()
    ap = w.leave("王师")
    w.assign(ap[0]["id"], type="substitute", handler_teacher_id=w.teachers["陈师"])

    d = _stats(w)["details"][0]
    assert d["handler_name"] == "陈师"
    assert d["absent_teacher_name"] == "王师"
    assert d["leave_type_label"] == "病假"
    assert d["sub_type_label"] == "代课"
    assert d["counts_toward_hours"] is True
    assert d["class_names"] == "701"
    assert d["subject_name"] == "语文"


def test_month_filter_excludes_other_months(w):
    w.teacher("王师", ["语文"])
    w.teacher("陈师", ["语文"])
    w.place("王师", "语文", "701", 0)
    w.publish()
    ap = w.leave("王师", when=WED)
    w.assign(ap[0]["id"], type="substitute", handler_teacher_id=w.teachers["陈师"])

    other_year, other_month = _other_month(WED)
    assert _stats(w)["summaries"]                                          # 当月有
    assert _stats(w, year=other_year, month=other_month)["summaries"] == []  # 别的月份没有


# ── 边界:跨月假单拆月计 ─────────────────────────────────────
def test_cross_month_leave_splits_by_period_date(w):
    """林师请的假横跨月底(含两个周三,分属前后两个月),各月各计一节。

    分月依据是「每个受影响节次自己的日期」,不是假单的起始月。
    """
    assert CROSS_WED.month != CROSS_WED2.month, "前置条件:两个周三必须跨月"
    w.teacher("林师", ["数学"])
    w.teacher("周师", ["数学"])
    w.place("林师", "数学", "701", 0, weekday=3)  # 每周三第一节
    w.publish()

    leave = w.client.post(f"/api/leaves{w.q}", json={
        "teacher_id": w.teachers["林师"], "leave_type": "official",
        "start_date": CROSS_WED.isoformat(), "end_date": CROSS_WED2.isoformat()}).json()
    aps = leave["affected_periods"]
    dates = sorted(p["date"] for p in aps)
    assert dates == [CROSS_WED.isoformat(), CROSS_WED2.isoformat()], dates
    for p in aps:
        w.assign(p["id"], type="substitute", handler_teacher_id=w.teachers["周师"])

    first = _stats(w, year=CROSS_WED.year, month=CROSS_WED.month)["summaries"]
    second = _stats(w, year=CROSS_WED2.year, month=CROSS_WED2.month)["summaries"]
    assert next(s for s in first if s["teacher_name"] == "周师")["billable_count"] == 1
    assert next(s for s in second if s["teacher_name"] == "周师")["billable_count"] == 1


# ── 销假的节次不计 ───────────────────────────────────────────
def test_cancelled_leave_excluded(w):
    w.teacher("王师", ["语文"])
    w.teacher("陈师", ["语文"])
    w.place("王师", "语文", "701", 0)
    w.publish()
    leave = w.client.post(f"/api/leaves{w.q}", json={
        "teacher_id": w.teachers["王师"], "leave_type": "sick",
        "start_date": WED.isoformat(), "end_date": WED.isoformat()}).json()
    w.assign(leave["affected_periods"][0]["id"],
             type="substitute", handler_teacher_id=w.teachers["陈师"])
    assert _stats(w)["summaries"], "销假前应计入"

    w.client.post(f"/api/leaves/{leave['id']}/cancel")
    assert _stats(w)["summaries"] == [], "销假后该节不计(那堂课没上)"


# ── 教师个人查询:只看自己 ──────────────────────────────────
def test_mine_shows_only_own(w):
    w.teacher("王师", ["语文"])
    w.teacher("陈师", ["语文"])
    w.teacher("周师", ["语文"])
    w.place("王师", "语文", "701", 0)
    w.place("王师", "语文", "702", 1)
    w.publish()
    ap = w.leave("王师")
    w.assign(ap[0]["id"], type="substitute", handler_teacher_id=w.teachers["陈师"])
    w.assign(ap[1]["id"], type="substitute", handler_teacher_id=w.teachers["周师"])

    # 绑定 chenacc 账号到陈师
    user = make_user(w.db, "chenacc", PW, roles=[Role.teacher])
    chen = w.db.get(Teacher, w.teachers["陈师"])
    chen.user_id = user.id
    w.db.commit()

    w.client.post("/api/auth/logout")
    w.client.post("/api/auth/login", json={"username": "chenacc", "password": PW})
    mine = w.client.get(
        f"/api/substitution-stats/mine{w.q}&year={WED.year}&month={WED.month}").json()
    names = {d["handler_name"] for d in mine["details"]}
    assert names == {"陈师"}, names
    assert len(mine["details"]) == 1


def test_teacher_cannot_view_full_stats(w):
    make_user(w.db, "t", PW, roles=[Role.teacher])
    w.client.post("/api/auth/logout")
    w.client.post("/api/auth/login", json={"username": "t", "password": PW})
    r = w.client.get(f"/api/substitution-stats{w.q}&year={WED.year}&month={WED.month}")
    assert r.status_code == 403


# ── Excel 导出 ───────────────────────────────────────────────
def test_export_xlsx(w):
    w.teacher("王师", ["语文"])
    w.teacher("陈师", ["语文"])
    w.place("王师", "语文", "701", 0)
    w.publish()
    ap = w.leave("王师")
    w.assign(ap[0]["id"], type="substitute", handler_teacher_id=w.teachers["陈师"])

    r = w.client.get(
        f"/api/substitution-stats/export{w.q}&year={WED.year}&month={WED.month}")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["汇总", "明细"]
    detail = wb["明细"]
    headers = [c.value for c in detail[1]]
    assert headers[:8] == [
        "教师",
        "日期",
        "节次",
        "班级",
        "科目",
        "原授课教师",
        "请假类型",
        "处理方式",
    ]
    assert "计费" in headers
    row = [c.value for c in detail[2]]
    assert row[0] == "陈师"
    assert row[5] == "王师"       # 原授课教师
    assert row[8] == "是"         # 计费
